"""report.py — 결과를 Excel/JSON/Markdown 으로 정리하고 최적 설정(참조 코드·모델 정보)을 산출.

산출물
------
1. report.xlsx  — Summary / Leaderboard / ConfusionMatrix / PerFormat / PerCategory /
                  Misclassified / LLM_vs_BERT (등급 비교 데이터)
2. optimized_config.json — 하네스 네이티브 최적 설정(전체)
3. weights.json          — data_classifier 공개 API(weights/model/ensemble)에 바로 쓰는 부분
4. OPTIMIZED_MODEL.md    — 모델 정보·KPI(기본 vs 최적 vs LLM)·성능·적용법(참조 코드)
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List, Optional

from . import metrics as M
from .db import DB

GRADES = ["O", "S", "C"]


# ════════════════════════════════════════════════════════════════════════
# 비교 데이터: LLM vs BERT(3-tier) vs 정답
# ════════════════════════════════════════════════════════════════════════
def build_comparison(db: DB, best_cid: str, split: str = "test") -> dict:
    """best 설정 예측 + LLM 예측 + 정답 조인 → 등급 비교 데이터 + 각 KPI."""
    bert = {(r["doc_id"], r["fmt"]): r for r in db.query(
        "SELECT * FROM predictions WHERE config_id=? AND split=?", (best_cid, split))}
    llm = {(r["doc_id"], r["fmt"]): r for r in db.query("SELECT * FROM llm_predictions")}
    rows = []
    for key, b in bert.items():
        l = llm.get(key)
        rows.append({
            "doc_id": key[0], "fmt": key[1], "true_grade": b["true_grade"],
            "bert_grade": b["pred_grade"], "bert_latency_ms": round(b["elapsed_ms"], 2),
            "llm_grade": (l["llm_grade"] if l else None),
            "llm_latency_ms": (round(l["latency_ms"], 1) if l else None),
            "bert_correct": b["pred_grade"] == b["true_grade"],
            "llm_correct": (l["llm_grade"] == b["true_grade"]) if l else None,
            "agree": (l["llm_grade"] == b["pred_grade"]) if l else None,
        })
    # KPI (LLM 예측이 있는 행만으로 공정 비교)
    paired = [r for r in rows if r["llm_grade"] is not None]
    bert_rows = [{"true_grade": r["true_grade"], "pred_grade": r["bert_grade"],
                  "fmt": r["fmt"], "latency_ms": r["bert_latency_ms"]} for r in paired]
    llm_rows = [{"true_grade": r["true_grade"], "pred_grade": r["llm_grade"],
                 "fmt": r["fmt"], "latency_ms": r["llm_latency_ms"]} for r in paired]
    agree = sum(1 for r in paired if r["agree"]) / len(paired) if paired else 0.0
    return {
        "rows": rows, "n_paired": len(paired),
        "agreement": round(agree, 4),
        "bert_metrics": M.compute(bert_rows) if bert_rows else {},
        "llm_metrics": M.compute(llm_rows) if llm_rows else {},
    }


# ════════════════════════════════════════════════════════════════════════
# Excel
# ════════════════════════════════════════════════════════════════════════
def export_excel(db: DB, run_id: str, best_cid: str, out_path: str,
                 default_cid: Optional[str], comparison: dict, perf: dict,
                 llm_summary: dict, best_cfg: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDDDDD")
    hl_fill = PatternFill("solid", fgColor="FFF3C4")   # XLSX/HWPX 강조

    def sheet(title):
        ws = wb.create_sheet(title)
        return ws

    def write_header(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = bold; c.fill = hdr_fill

    run = db.query("SELECT * FROM runs WHERE run_id=?", (run_id,))
    run = run[0] if run else {}
    best_metrics = db.query("SELECT detail FROM metrics WHERE config_id=? AND split='test'", (best_cid,))
    best_test = DB.jload(best_metrics[0]["detail"]) if best_metrics else {}
    def_test = {}
    if default_cid:
        dm = db.query("SELECT detail FROM metrics WHERE config_id=? AND split='test'", (default_cid,))
        def_test = DB.jload(dm[0]["detail"]) if dm else {}

    # ── Summary ──
    ws = wb.active; ws.title = "Summary"
    ws.append(["data_classifier 자동 평가·최적화 결과"]); ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["run_id", run.get("run_id")])
    ws.append(["코퍼스", f"문서 {run.get('n_docs')}개 × 포맷 → 파일 {run.get('n_files')}개 (seed={run.get('corpus_seed')})"])
    ws.append([])
    ws.append(["최적 설정 (BEST)", json.dumps(best_cfg, ensure_ascii=False)])
    ws.append([])
    write_at = ws.max_row + 1
    ws.append(["KPI (test split)", "기본(default)", "최적(BEST)", "LLM(Claude)"])
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = hdr_fill
    lm = comparison.get("llm_metrics", {})
    bm = comparison.get("bert_metrics", {})
    def row(name, key, fmt="{:.4f}"):
        ws.append([name,
                   fmt.format(def_test.get(key, 0)) if def_test else "-",
                   fmt.format(best_test.get(key, 0)) if best_test else "-",
                   fmt.format(lm.get(key, 0)) if lm else "-"])
    row("macro_f1", "macro_f1"); row("accuracy", "accuracy")
    row("C_recall(기밀 재현율)", "c_recall"); row("C_precision", "c_precision")
    row("under_rate(과소분류=유출)", "under_rate"); row("over_rate(과대분류)", "over_rate")
    ws.append([])
    ws.append(["성능(속도)", "BERT 3-tier", "LLM(Claude)"]);
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = hdr_fill
    ws.append(["end-to-end p50 (ms)", perf.get("end_to_end_p50_ms"), llm_summary.get("latency_p50_ms")])
    ws.append(["end-to-end p95 (ms)", perf.get("end_to_end_p95_ms"), llm_summary.get("latency_p95_ms")])
    ws.append(["뉴럴 추론(ms, 포함시)", perf.get("neural_infer_ms"), "-"])
    ws.append([])
    ws.append(["LLM-BERT 일치율(agreement)", comparison.get("agreement")])
    ws.append(["LLM 토큰 in/out", f"{llm_summary.get('input_tokens','-')}/{llm_summary.get('output_tokens','-')}"])
    ws.append(["참고", "LLM 은 깨끗한 추출 텍스트를 입력받음. 3-tier 모델은 추출+분류 모두 수행(비대칭)."])

    # ── Leaderboard ──
    ws = sheet("Leaderboard")
    write_header(ws, ["config_id", "kind", "label", "valid_obj", "valid_macroF1",
                      "valid_Crecall", "valid_under", "valid_over",
                      "test_macroF1", "test_acc", "test_Crecall", "params"])
    lb = db.query("""SELECT c.config_id, c.kind, c.label, c.params,
                            v.objective vo, v.macro_f1 vf, v.c_recall vc, v.under_rate vu, v.over_rate ov,
                            t.macro_f1 tf, t.accuracy ta, t.c_recall tc
                     FROM configs c
                     LEFT JOIN metrics v ON v.config_id=c.config_id AND v.split='valid'
                     LEFT JOIN metrics t ON t.config_id=c.config_id AND t.split='test'
                     WHERE c.run_id=? ORDER BY v.objective DESC""", (run_id,))
    for r in lb:
        ws.append([r["config_id"], r["kind"], r["label"], r["vo"], r["vf"], r["vc"],
                   r["vu"], r["ov"], r["tf"], r["ta"], r["tc"], r["params"]])

    # ── ConfusionMatrix (best, test) ──
    ws = sheet("ConfusionMatrix")
    ws.append(["최적 설정 혼동행렬 (test) — 행=실제, 열=예측"]); ws["A1"].font = bold
    conf = best_test.get("confusion", {})
    write_header_at = 3
    ws.append([])
    ws.append(["", "→O", "→S", "→C"])
    for c in ws[ws.max_row]: c.font = bold
    for t in GRADES:
        ws.append([f"실제 {t}", conf.get(t, {}).get("O", 0), conf.get(t, {}).get("S", 0), conf.get(t, {}).get("C", 0)])

    # ── PerFormat (XLSX/HWPX 강조) ──
    ws = sheet("PerFormat")
    ws.append(["포맷별 정확도 (최적 설정, test) — 추출 레이어 결함 분리"]); ws["A1"].font = bold
    ws.append([])
    write_header(ws, ["format", "n", "accuracy", "under_rate", "over_rate"]) if False else None
    ws.append(["format", "n", "accuracy", "under_rate", "over_rate"])
    for c in ws[ws.max_row]: c.font = bold; c.fill = hdr_fill
    bf = best_test.get("by_format", {})
    for fmt in sorted(bf):
        ws.append([fmt, bf[fmt]["n"], bf[fmt]["accuracy"], bf[fmt]["under_rate"], bf[fmt]["over_rate"]])
        if fmt in ("xlsx", "hwpx"):
            for c in ws[ws.max_row]:
                c.fill = hl_fill

    # ── PerCategory ──
    ws = sheet("PerCategory")
    ws.append(["난이도별 정확도 (최적 설정, test)"]); ws["A1"].font = bold
    ws.append([])
    ws.append(["category", "n", "accuracy", "under_rate", "over_rate"])
    for c in ws[ws.max_row]: c.font = bold; c.fill = hdr_fill
    bc = best_test.get("by_category", {})
    for cat in sorted(bc):
        ws.append([cat, bc[cat]["n"], bc[cat]["accuracy"], bc[cat]["under_rate"], bc[cat]["over_rate"]])

    # ── Misclassified (best, test) ──
    ws = sheet("Misclassified")
    write_header(ws, ["doc_id", "fmt", "true", "pred(BERT)", "score", "llm"])
    cmp_by = {(r["doc_id"], r["fmt"]): r for r in comparison["rows"]}
    for r in db.query("SELECT * FROM predictions WHERE config_id=? AND split='test'", (best_cid,)):
        if r["pred_grade"] != r["true_grade"]:
            c = cmp_by.get((r["doc_id"], r["fmt"]), {})
            ws.append([r["doc_id"], r["fmt"], r["true_grade"], r["pred_grade"],
                       round(r["score"], 2), c.get("llm_grade")])

    # ── LLM_vs_BERT 비교 데이터 ──
    ws = sheet("LLM_vs_BERT")
    write_header(ws, ["doc_id", "fmt", "true", "BERT", "LLM", "BERT_ok", "LLM_ok",
                      "agree", "BERT_ms", "LLM_ms"])
    for r in comparison["rows"]:
        ws.append([r["doc_id"], r["fmt"], r["true_grade"], r["bert_grade"], r["llm_grade"],
                   r["bert_correct"], r["llm_correct"], r["agree"],
                   r["bert_latency_ms"], r["llm_latency_ms"]])

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ════════════════════════════════════════════════════════════════════════
# 최적 설정 산출물 (참조 코드 · 모델 정보)
# ════════════════════════════════════════════════════════════════════════
def export_config_files(out_dir: Path, best_cfg: dict, best_test: dict, def_test: dict,
                        comparison: dict, perf: dict, llm_summary: dict,
                        models_info: dict):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) 하네스 네이티브 전체 설정
    (out_dir / "optimized_config.json").write_text(
        json.dumps(best_cfg, ensure_ascii=False, indent=2), encoding="utf-8")

    # 2) data_classifier 공개 API 용 weights.json (entity/keyword/thresholds/tier)
    weights = {k: best_cfg[k] for k in ("entity", "keyword", "thresholds", "tier") if k in best_cfg}
    (out_dir / "weights.json").write_text(
        json.dumps(weights, ensure_ascii=False, indent=2), encoding="utf-8")

    # 3) 모델 정보 + 적용법 Markdown
    nb = best_cfg.get("model")
    neural_meta = models_info.get(nb, {}) if nb else {}
    sup = best_cfg.get("supersede")
    lm = comparison.get("llm_metrics", {})
    bm = best_test

    def fmt(d, k):
        return f"{d.get(k, 0):.4f}" if d else "-"

    cli = f"python data_classifier.py <file> --locale ko --weights weights.json"
    if best_cfg.get("llm_mode"):
        cli += f" --llm --model {nb} --ensemble {best_cfg.get('ensemble', 'escalate')}"

    md = f"""# 최적화된 등급 분류 모델 — 참조 설정 · 모델 정보

자동 탐색으로 선정된 data_classifier.py 최적 설정과 성능 근거를 정리한다.

## 1. 모델 구성 (Model Info)

| 항목 | 값 |
|---|---|
| 신경망(BERT) 백엔드 | {('`'+nb+'` — '+neural_meta.get('label','')) if nb else '미사용(규칙+NER 전용)'} |
| 백엔드 종류 | {neural_meta.get('kind','-') if nb else '-'} |
| HF 모델 | {('`'+neural_meta.get('model','')+'`') if neural_meta else '-'} |
| 앙상블 방식 | `{best_cfg.get('ensemble','escalate') if best_cfg.get('llm_mode') else 'escalate(뉴럴 미사용)'}` |
| C(기밀) 임계값 | {best_cfg.get('thresholds',{}).get('confidential', 5.5)} |
| S(민감) 임계값 | {best_cfg.get('thresholds',{}).get('sensitive', 0.75)} |
| 엔티티 가중치 override | `{json.dumps(best_cfg.get('entity',{}), ensure_ascii=False)}` |
| tier 가중치 | `{json.dumps(best_cfg.get('tier',{}), ensure_ascii=False) or '기본'}` |
| supersede 규칙 | `{json.dumps(sup, ensure_ascii=False) if sup else '없음'}` |

## 2. 성능 KPI (test split, 정답=정책 라벨)

| 지표 | 기본 설정 | **최적 설정(BERT 3-tier)** | LLM(Claude {llm_summary.get('model','-')}) |
|---|---|---|---|
| macro F1 | {fmt(def_test,'macro_f1')} | **{fmt(bm,'macro_f1')}** | {fmt(lm,'macro_f1')} |
| accuracy | {fmt(def_test,'accuracy')} | **{fmt(bm,'accuracy')}** | {fmt(lm,'accuracy')} |
| C 재현율(기밀 누락 방지) | {fmt(def_test,'c_recall')} | **{fmt(bm,'c_recall')}** | {fmt(lm,'c_recall')} |
| C 정밀도 | {fmt(def_test,'c_precision')} | **{fmt(bm,'c_precision')}** | {fmt(lm,'c_precision')} |
| 과소분류율(유출 위험) | {fmt(def_test,'under_rate')} | **{fmt(bm,'under_rate')}** | {fmt(lm,'under_rate')} |
| 과대분류율 | {fmt(def_test,'over_rate')} | **{fmt(bm,'over_rate')}** | {fmt(lm,'over_rate')} |

LLM 대비 일치율(agreement): **{comparison.get('agreement','-')}**  (비교 파일 {comparison.get('n_paired','-')}건)

## 3. 성능(속도) — GPU 없이 LLM 대비

| | BERT 3-tier(CPU) | LLM(Claude) |
|---|---|---|
| end-to-end p50 | {perf.get('end_to_end_p50_ms','-')} ms | {llm_summary.get('latency_p50_ms','-')} ms |
| end-to-end p95 | {perf.get('end_to_end_p95_ms','-')} ms | {llm_summary.get('latency_p95_ms','-')} ms |

> LLM 은 깨끗한 추출 텍스트를 입력받고, 3-tier 모델은 추출+분류를 모두 수행한다(비대칭). 그럼에도
> 외부 LLM/GPU 없이 위 KPI/속도를 달성한다.

## 4. 적용법 (참조 코드)

### CLI
```bash
{cli}
```

### Python
```python
import json
from data_classifier import classify
weights = json.load(open("weights.json", encoding="utf-8"))
r = classify("문서.xlsx", n2sf_mode=True, locale="ko",
             weights=weights,{(' llm_mode=True, model="'+nb+'", ensemble_method="'+best_cfg.get('ensemble','escalate')+'",') if best_cfg.get('llm_mode') else ''}
             )
print(r["grade"], r["score"])
```
"""
    if sup:
        # supersede 는 공개 API 미지원 → SUPERSEDED_BY 1줄 추가 패치 안내
        loser = list(sup.keys())[0]
        winners = sup[loser]
        md += f"""
### 5. 탐지 레이어 패치 (supersede — 권장)

최적 설정의 `supersede={json.dumps(sup, ensure_ascii=False)}` 는 공개 `weights` API 로 표현되지
않는다(탐지 단계 규칙). data_classifier.py 의 `SUPERSEDED_BY` 에 아래 한 줄을 추가하면
동일 효과(겹치는 일반 인식기를 더 구체적인 인식기가 억제)를 내부적으로 얻는다:

```python
SUPERSEDED_BY = {{
    ...,
    "{loser}": {json.dumps(winners, ensure_ascii=False)},   # 전화번호와 겹친 계좌 오탐 제거
}}
```
이 패치는 전화번호가 느슨한 계좌 정규식에 오탐되어 S→C 로 과대분류되던 문제를 교정한다.
"""
    (out_dir / "OPTIMIZED_MODEL.md").write_text(md, encoding="utf-8")
    return {"optimized_config": str(out_dir / "optimized_config.json"),
            "weights": str(out_dir / "weights.json"),
            "model_md": str(out_dir / "OPTIMIZED_MODEL.md")}
